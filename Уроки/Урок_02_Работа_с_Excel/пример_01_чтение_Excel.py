"""
Пример 2.1: Базовое чтение данных из Excel

Демонстрирует основные способы чтения данных из Excel файлов
"""

from openpyxl import load_workbook
from pathlib import Path


def пример_чтения_ячеек():
    """Чтение конкретных ячеек из Excel"""
    print("=" * 60)
    print("ПРИМЕР 1: Чтение конкретных ячеек")
    print("=" * 60)
    
    # Создаем тестовый файл для демонстрации
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Параметры"
    
    # Заполняем тестовыми данными
    ws["A1"] = "Параметр"
    ws["B1"] = "Значение"
    ws["A2"] = "Напряжение"
    ws["B2"] = 220.5
    ws["A3"] = "Мощность"
    ws["B3"] = 1000.0
    ws["A4"] = "Режим"
    ws["B4"] = "установившийся"
    
    # Сохраняем временный файл
    тестовый_файл = "тест_данные.xlsx"
    wb.save(тестовый_файл)
    print(f"✅ Создан тестовый файл: {тестовый_файл}\n")
    
    # Теперь читаем данные
    workbook = load_workbook(тестовый_файл, data_only=True)
    worksheet = workbook["Параметры"]
    
    # Способ 1: Чтение по адресу ячейки
    print("Способ 1: Чтение по адресу ячейки (например, 'B2'):")
    напряжение = worksheet["B2"].value
    print(f"  Напряжение: {напряжение} В")
    
    # Способ 2: Чтение через cell(row, column)
    print("\nСпособ 2: Чтение через cell(row, column):")
    мощность = worksheet.cell(row=3, column=2).value
    print(f"  Мощность: {мощность} МВт")
    
    # Способ 3: Чтение с проверкой типа
    режим = worksheet["B4"].value
    print(f"\nРежим: {режим} (тип: {type(режим).__name__})")
    
    workbook.close()
    
    # Удаляем тестовый файл
    Path(тестовый_файл).unlink()
    print(f"\n✅ Тестовый файл удален")


def пример_чтения_таблицы():
    """Чтение таблицы построчно"""
    print("\n" + "=" * 60)
    print("ПРИМЕР 2: Чтение таблицы построчно")
    print("=" * 60)
    
    # Создаем тестовую таблицу
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Генераторы"
    
    # Заголовки
    заголовки = ["Генератор", "P_исх, МВт", "P_max, МВт", "P_min, МВт"]
    for col_idx, заголовок in enumerate(заголовки, start=1):
        ws.cell(row=1, column=col_idx, value=заголовок)
    
    # Данные
    данные_генераторов = [
        ["Г-1", 150, 200, 50],
        ["Г-2", 200, 250, 60],
        ["Г-3", 180, 220, 55],
    ]
    
    for row_idx, строка_данных in enumerate(данные_генераторов, start=2):
        for col_idx, значение in enumerate(строка_данных, start=1):
            ws.cell(row=row_idx, column=col_idx, value=значение)
    
    тестовый_файл = "тест_генераторы.xlsx"
    wb.save(тестовый_файл)
    print(f"✅ Создан тестовый файл: {тестовый_файл}\n")
    
    # Читаем таблицу
    workbook = load_workbook(тестовый_файл, data_only=True)
    worksheet = workbook["Генераторы"]
    
    # Читаем заголовки (первая строка)
    заголовки = []
    for cell in worksheet[1]:
        заголовки.append(cell.value)
    print("Заголовки:", заголовки)
    
    # Читаем данные (со второй строки)
    print("\nДанные генераторов:")
    генераторы = []
    
    for row_idx in range(2, worksheet.max_row + 1):
        генератор = {}
        for col_idx, заголовок in enumerate(заголовки, start=1):
            значение = worksheet.cell(row=row_idx, column=col_idx).value
            генератор[заголовок] = значение
        
        генераторы.append(генератор)
        print(f"  {генератор}")
    
    workbook.close()
    Path(тестовый_файл).unlink()
    print(f"\n✅ Тестовый файл удален")


def пример_чтения_с_поиском():
    """Чтение данных с поиском по ключевым словам"""
    print("\n" + "=" * 60)
    print("ПРИМЕР 3: Чтение с поиском по ключевым словам")
    print("=" * 60)
    
    # Создаем файл с параметрами
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    
    параметры = {
        "Напряжение сети": 220.5,
        "Мощность": 1000.0,
        "Время расчета": 5.0,
        "Режим": "установившийся",
    }
    
    # Записываем в две колонки
    row = 1
    for ключ, значение in параметры.items():
        ws.cell(row=row, column=1, value=ключ)
        ws.cell(row=row, column=2, value=значение)
        row += 1
    
    тестовый_файл = "тест_параметры.xlsx"
    wb.save(тестовый_файл)
    print(f"✅ Создан тестовый файл: {тестовый_файл}\n")
    
    # Читаем с поиском
    workbook = load_workbook(тестовый_файл, data_only=True)
    worksheet = workbook.active
    
    извлеченные = {}
    
    # Ищем нужные параметры
    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row):
        ключ = row[0].value
        значение = row[1].value
        
        if ключ and isinstance(ключ, str):
            # Ищем по ключевым словам
            if "Напряжение" in ключ:
                извлеченные["напряжение"] = значение
            elif "Мощность" in ключ:
                извлеченные["мощность"] = значение
            elif "Время" in ключ:
                извлеченные["время"] = значение
    
    print("Извлеченные параметры:")
    for ключ, значение in извлеченные.items():
        print(f"  {ключ}: {значение}")
    
    workbook.close()
    Path(тестовый_файл).unlink()
    print(f"\n✅ Тестовый файл удален")


# Запуск примеров
if __name__ == "__main__":
    пример_чтения_ячеек()
    пример_чтения_таблицы()
    пример_чтения_с_поиском()
    
    print("\n" + "=" * 60)
    print("✅ ВСЕ ПРИМЕРЫ ЗАВЕРШЕНЫ")
    print("=" * 60)

