"""
ПРИМЕР: Запись результатов расчета в Excel

Этот пример показывает, как записывать результаты расчетов
динамической устойчивости в Excel файл.
"""

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from pathlib import Path


def создать_шаблон_результатов(путь_к_файлу):
    """
    Создает шаблон Excel файла для результатов
    
    Args:
        путь_к_файлу: Путь для сохранения файла
    """
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Результаты"
    
    # Заголовки
    заголовки = [
        "Генератор",
        "P_исх, МВт",
        "P_уст, МВт",
        "P_неуст, МВт",
        "Состояние",
    ]
    
    # Записываем заголовки
    for col_idx, заголовок in enumerate(заголовки, start=1):
        cell = worksheet.cell(1, col_idx)
        cell.value = заголовок
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Устанавливаем ширину столбцов
    worksheet.column_dimensions['A'].width = 20
    worksheet.column_dimensions['B'].width = 15
    worksheet.column_dimensions['C'].width = 15
    worksheet.column_dimensions['D'].width = 15
    worksheet.column_dimensions['E'].width = 15
    
    workbook.save(путь_к_файлу)
    print(f"✅ Создан шаблон: {путь_к_файлу}")


def записать_результаты_генераторов(путь_к_файлу, результаты_генераторов):
    """
    Записывает результаты генераторов в Excel
    
    Args:
        путь_к_файлу: Путь к Excel файлу
        результаты_генераторов: Список словарей с результатами
        [
            {
                "Название": "Г-1",
                "P_исх": 150.0,
                "P_уст": 145.0,
                "P_неуст": None,
                "Состояние": "устойчиво"
            },
            ...
        ]
    """
    if not Path(путь_к_файлу).exists():
        создать_шаблон_результатов(путь_к_файлу)
    
    workbook = load_workbook(путь_к_файлу)
    
    # Получаем или создаем лист
    if "Результаты" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("Результаты")
    else:
        worksheet = workbook["Результаты"]
    
    # Начинаем запись со второй строки (первая - заголовки)
    начальная_строка = 2
    
    for row_idx, генератор in enumerate(результаты_генераторов, start=начальная_строка):
        # Записываем данные
        worksheet.cell(row_idx, 1, генератор.get("Название", ""))
        worksheet.cell(row_idx, 2, генератор.get("P_исх", 0))
        worksheet.cell(row_idx, 3, генератор.get("P_уст") or "")
        worksheet.cell(row_idx, 4, генератор.get("P_неуст") or "")
        worksheet.cell(row_idx, 5, генератор.get("Состояние", ""))
        
        # Форматирование
        for col in range(1, 6):
            cell = worksheet.cell(row_idx, col)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Цветовая индикация для состояния
            if col == 5:  # Столбец "Состояние"
                состояние = генератор.get("Состояние", "")
                if "устойчиво" in состояние.lower():
                    cell.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                elif "неустойчиво" in состояние.lower():
                    cell.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
    
    workbook.save(путь_к_файлу)
    print(f"✅ Результаты записаны в {путь_к_файлу}")


def записать_общие_результаты(путь_к_файлу, общие_результаты):
    """
    Записывает общие результаты расчета в Excel
    
    Args:
        путь_к_файлу: Путь к Excel файлу
        общие_результаты: Словарь с общими результатами
        {
            "Вид расчета": "Проверка ДУ",
            "Система устойчива": True,
            "Время расчета": "0:05:30",
            "Количество итераций": 15,
            ...
        }
    """
    if not Path(путь_к_файлу).exists():
        workbook = Workbook()
        workbook.active.title = "Общие результаты"
        workbook.save(путь_к_файлу)
    
    workbook = load_workbook(путь_к_файлу)
    
    # Получаем или создаем лист
    if "Общие результаты" not in workbook.sheetnames:
        worksheet = workbook.create_sheet("Общие результаты")
    else:
        worksheet = workbook["Общие результаты"]
    
    # Записываем результаты построчно
    row = 1
    for ключ, значение in общие_результаты.items():
        worksheet.cell(row, 1, ключ)
        worksheet.cell(row, 2, str(значение))
        
        # Форматирование заголовков
        cell_ключ = worksheet.cell(row, 1)
        cell_ключ.font = Font(bold=True)
        
        # Цветовая индикация для результата устойчивости
        if "устойчива" in ключ.lower() and isinstance(значение, bool):
            cell_значение = worksheet.cell(row, 2)
            if значение:
                cell_значение.fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
            else:
                cell_значение.fill = PatternFill(start_color="FF6347", end_color="FF6347", fill_type="solid")
        
        row += 1
    
    # Устанавливаем ширину столбцов
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 20
    
    workbook.save(путь_к_файлу)
    print(f"✅ Общие результаты записаны в {путь_к_файлу}")


def записать_результаты_из_словаря(путь_к_файлу, словарь_результатов, лист="Результаты"):
    """
    Записывает результаты из словаря в Excel (как в реальных проектах)
    
    Это более универсальная функция, которая работает со сложными структурами
    
    Args:
        путь_к_файлу: Путь к Excel файлу
        словарь_результатов: Словарь с результатами (может быть вложенным)
        лист: Название листа
    """
    if not Path(путь_к_файлу).exists():
        workbook = Workbook()
        workbook.active.title = лист
        workbook.save(путь_к_файлу)
    
    workbook = load_workbook(путь_к_файлу)
    
    if лист not in workbook.sheetnames:
        worksheet = workbook.create_sheet(лист)
    else:
        worksheet = workbook[лист]
    
    row = 1
    
    def записать_вложенный_словарь(данные, уровень=0, row_start=1):
        """Рекурсивная функция для записи вложенных словарей"""
        текущая_строка = row_start
        
        for ключ, значение in данные.items():
            # Записываем ключ с отступом
            отступ = "  " * уровень
            worksheet.cell(текущая_строка, 1 + уровень, отступ + str(ключ))
            
            # Если значение - словарь, рекурсивно обрабатываем
            if isinstance(значение, dict):
                worksheet.cell(текущая_строка, 1 + уровень).font = Font(bold=True)
                текущая_строка = записать_вложенный_словарь(значение, уровень + 1, текущая_строка + 1)
            else:
                # Записываем значение
                col = 1 + уровень + 1
                worksheet.cell(текущая_строка, col, str(значение))
                текущая_строка += 1
        
        return текущая_строка
    
    # Записываем данные
    row = записать_вложенный_словарь(словарь_результатов, row_start=1)
    
    workbook.save(путь_к_файлу)
    print(f"✅ Результаты из словаря записаны в {путь_к_файлу}")


# Пример использования
if __name__ == "__main__":
    print("="*60)
    print("ПРИМЕР: Запись результатов в Excel")
    print("="*60)
    
    путь = "результаты_расчета.xlsx"
    
    # Пример 1: Результаты генераторов
    print("\n1. Запись результатов генераторов:")
    результаты_генераторов = [
        {
            "Название": "Г-1",
            "P_исх": 150.0,
            "P_уст": 145.0,
            "P_неуст": None,
            "Состояние": "устойчиво"
        },
        {
            "Название": "Г-2",
            "P_исх": 200.0,
            "P_уст": None,
            "P_неуст": 210.0,
            "Состояние": "неустойчиво"
        },
    ]
    записать_результаты_генераторов(путь, результаты_генераторов)
    
    # Пример 2: Общие результаты
    print("\n2. Запись общих результатов:")
    общие_результаты = {
        "Вид расчета": "Проверка ДУ",
        "Система устойчива": True,
        "Время расчета": "0:05:30",
        "Количество итераций": 15,
        "Превышение угла": "-",
    }
    записать_общие_результаты(путь, общие_результаты)
    
    # Пример 3: Сложная структура (как в реальных проектах)
    print("\n3. Запись сложной структуры результатов:")
    сложные_результаты = {
        "id": 1,
        "Вид расчета": "Проверка ДУ",
        "Генераторы": {
            "Г-1": {
                "P_уст": 145.0,
                "P_неуст": None,
            },
            "Г-2": {
                "P_уст": None,
                "P_неуст": 210.0,
            },
        },
        "Результаты расчета динамики": {
            "Система устойчива": True,
            "Время достигнутое": 5.0,
        },
    }
    записать_результаты_из_словаря(путь, сложные_результаты, "Детальные результаты")
    
    print(f"\n✅ Все результаты записаны в {путь}")
    print("   Откройте файл в Excel для просмотра")

